import os
import pytest
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.excel_parser import parse_excel_file, FormulaInfo
from core.csv_tsv_parser import parse_csv_tsv_file
from core.formula_extractor import extract_formulas_from_file, _sanitize_header, extract_and_sanitize_headers

# Helper to create a dummy Excel file
@pytest.fixture
def excel_file(tmp_path):
    file_path = tmp_path / "test.xlsx"
    workbook = Workbook()
    sheet: Worksheet = workbook.active # type: ignore
    assert sheet is not None, "Active sheet should not be None"
    sheet.title = "Sheet1"
    sheet['A1'] = "=1+1"
    sheet['B2'] = "=SUM(A1,1)"
    sheet2: Worksheet = workbook.create_sheet("Sheet2") # type: ignore
    sheet2['A1'] = "=Sheet1!A1+Sheet1!B2"
    workbook.save(file_path)
    return file_path

# Helper to create a dummy CSV file
@pytest.fixture
def csv_file(tmp_path):
    file_path = tmp_path / "test.csv"
    content = '"=1+1","Data1"\n"=SUM(A1,B1)","Data2"'
    file_path.write_text(content)
    return file_path

# Helper to create a dummy TSV file
@pytest.fixture
def tsv_file(tmp_path):
    file_path = tmp_path / "test.tsv"
    content = "=1+1\tData1\n=SUM(A1,B1)\tData2"
    file_path.write_text(content)
    return file_path

def test_parse_excel_file_basic(excel_file):
    formulas = parse_excel_file(excel_file)
    assert len(formulas) == 3
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" and f.sheet_name == "Sheet1" for f in formulas)
    assert any(f.formula_string == "=SUM(A1,1)" and f.cell_address == "B2" and f.sheet_name == "Sheet1" for f in formulas)
    assert any(f.formula_string == "=Sheet1!A1+Sheet1!B2" and f.cell_address == "A1" and f.sheet_name == "Sheet2" for f in formulas)

def test_parse_csv_file_basic(csv_file):
    formulas = parse_csv_tsv_file(csv_file)
    assert len(formulas) == 2
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" for f in formulas)
    assert any(f.formula_string == "=SUM(A1,B1)" and f.cell_address == "A2" for f in formulas) # Adjusted for 0-indexed row

def test_parse_tsv_file_basic(tsv_file):
    formulas = parse_csv_tsv_file(tsv_file)
    assert len(formulas) == 2
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" for f in formulas)
    assert any(f.formula_string == "=SUM(A1,B1)" and f.cell_address == "A2" for f in formulas) # Adjusted for 0-indexed row

def test_extract_formulas_excel(excel_file):
    formulas = extract_formulas_from_file(excel_file)
    assert len(formulas) == 3
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" and f.sheet_name == "Sheet1" for f in formulas)

def test_extract_formulas_csv(csv_file):
    formulas = extract_formulas_from_file(csv_file)
    assert len(formulas) == 2
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" for f in formulas)

def test_extract_formulas_tsv(tsv_file):
    formulas = extract_formulas_from_file(tsv_file)
    assert len(formulas) == 2
    assert any(f.formula_string == "=1+1" and f.cell_address == "A1" for f in formulas)

def test_extract_formulas_unsupported_type(tmp_path):
    unsupported_file = tmp_path / "test.txt"
    unsupported_file.write_text("just text")
    formulas = extract_formulas_from_file(unsupported_file)
    assert len(formulas) == 1
    assert formulas[0].error == f"Unsupported file type: .txt"

def test_sanitize_header():
    assert _sanitize_header("Revenue Q2") == "revenue_q2"
    assert _sanitize_header("Product-ID") == "product_id"
    assert _sanitize_header("  Total Cost ") == "total_cost"
    assert _sanitize_header("123_Header") == "col_123_header"
    assert _sanitize_header("!@#Special Chars") == "special_chars"
    assert _sanitize_header("") == ""

@pytest.fixture
def excel_file_with_headers(tmp_path):
    file_path = tmp_path / "headers.xlsx"
    workbook = Workbook()
    sheet: Worksheet = workbook.active # type: ignore
    assert sheet is not None, "Active sheet should not be None"
    sheet.title = "HeaderTest"
    
    # Simple headers
    sheet['A1'] = "Col A"
    sheet['B1'] = "Col B-Value"
    # Merged header (A2:B2 merged, value in A2)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    sheet['A2'] = "Merged Header"
    # Empty header
    sheet['C1'] = ""
    # Header with special characters
    sheet['D1'] = "Data #ID%"
    # Header starting with a number
    sheet['E1'] = "1st Quarter"
    workbook.save(file_path)
    return file_path

def test_extract_and_sanitize_headers(excel_file_with_headers):
    headers = extract_and_sanitize_headers(excel_file_with_headers)
    assert "HeaderTest" in headers
    sanitized_headers = headers["HeaderTest"]
    assert len(sanitized_headers) == 5 # Columns A, B, C, D, E
    assert sanitized_headers[0] == "col_a"
    assert sanitized_headers[1] == "col_b_value"
    assert sanitized_headers[2] == ""
    assert sanitized_headers[3] == "data_id"
    assert sanitized_headers[4] == "col_1st_quarter" 