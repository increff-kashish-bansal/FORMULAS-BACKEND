import os
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell # Explicitly import Cell type
from openpyxl.cell.cell import MergedCell # Also import MergedCell
from formulas import ExcelModel
from typing import Dict, List, Any
import re # Import re module for regex operations
from utils.logger import AppLogger # Import the custom logger

# Instantiate the custom logger
app_logger = AppLogger(__name__)

def get_excel_column_letter(col_idx: int) -> str:
    """Converts a 0-indexed column integer to an Excel column letter (e.g., 0 -> A, 26 -> AA)."""
    letter = ""
    while col_idx >= 0:
        letter = chr(65 + (col_idx % 26)) + letter
        col_idx = (col_idx // 26) - 1
    return letter

def extract_named_ranges(file_path: str) -> Dict[str, str]:
    """
    Extracts named ranges from an Excel file.

    Args:
        file_path (str): The path to the Excel file.

    Returns:
        Dict[str, str]: A dictionary where keys are named range names and values are their
                        corresponding cell addresses (e.g., {'MyRange': 'Sheet1!$A$1:$B$2'}).
    """
    named_ranges: Dict[str, str] = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False) # Changed to False to get formulas
        
        for defined_name in workbook.defined_names.definedName:
            # Defined names can refer to cells, ranges, or even constants. 
            # We are primarily interested in cell/range references.
            if defined_name.type == 'rng': # Check if it's a range type
                # Check if the named range refers to a dynamic formula (e.g., using OFFSET or INDEX)
                if defined_name.value and (re.search(r'OFFSET\(', defined_name.value, re.IGNORECASE) or \
                                           re.search(r'INDEX\(', defined_name.value, re.IGNORECASE)):
                    app_logger.warning(
                        f"Found dynamic named range '{defined_name.name}' with formula: {defined_name.value}. "
                        "The 'formulas' library is expected to handle its evaluation.",
                        context={'named_range': defined_name.name, 'formula': defined_name.value, 'file': file_path}
                    )
                    # We still store its value (the formula string) for potential later use or debugging
                    named_ranges[defined_name.name] = defined_name.value
                elif defined_name.destinations:
                    sheet_name, cell_address = list(defined_name.destinations)[0]
                    # The cell_address usually comes with sheet name prefixed if it's a sheet-specific named range
                    # but for global named ranges, it might just be the address.
                    # We'll store it in the format "SheetName!CellAddress"
                    full_address = f"{sheet_name}!{cell_address}" if sheet_name else cell_address
                    app_logger.info(
                        f"Found static named range: {defined_name.name} -> {full_address}",
                        context={'named_range': defined_name.name, 'address': full_address, 'file': file_path}
                    )
            elif defined_name.type == 'const':
                app_logger.info(
                    f"Skipping named constant: {defined_name.name} = {defined_name.value}",
                    context={'named_constant': defined_name.name, 'value': defined_name.value, 'file': file_path}
                )
            else:
                app_logger.info(
                    f"Skipping unknown defined name type '{defined_name.type}': {defined_name.name}",
                    context={'name': defined_name.name, 'type': defined_name.type, 'file': file_path}
                )

    except Exception as e:
        app_logger.error(f"Error extracting named ranges from {file_path}: {e}", context={'file': file_path, 'error': str(e)})

    return named_ranges

def parse_excel_file(file_path: str) -> list[Dict[str, str]]:
    """
    Parses an Excel file (.xls or .xlsx) to identify and extract formulas.
    Args:
        file_path (str): The path to the Excel file.
    Returns:
        list[Dict[str, str]]: A list of dictionaries containing extracted raw formula information.
    """
    extracted_formulas_raw = []
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        for sheet_name in workbook.sheetnames:
            sheet: Worksheet = workbook[sheet_name]
            app_logger.info(
                f"Processing sheet: {sheet_name} in {os.path.basename(file_path)}",
                context={'sheet': sheet_name, 'file': os.path.basename(file_path)}
            )

            # Iterate through all cells in the sheet
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)

                    # If it's a merged cell, we only process the top-left cell of the merged range
                    # The formula (if any) is typically stored in the top-left cell.
                    if isinstance(cell, MergedCell):
                        # We need to find the top-left cell of the merged range
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                # Check if the current cell is the top-left cell of the merged range
                                if cell.coordinate == merged_range.start_cell.coordinate:
                                    # Get the top-left cell which holds the value/formula for the merged range
                                    cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                                    app_logger.info(
                                        f"Processing merged cell range: {merged_range.coord} via its top-left cell: {cell.coordinate}",
                                        context={'range': merged_range.coord, 'top_left_cell': cell.coordinate, 'sheet': sheet_name}
                                    )
                                else:
                                    # Skip other cells within a merged range, as they don't hold the formula
                                    continue # Skip to the next cell in the loop

                    if isinstance(cell, Cell) and cell.data_type == 'f':
                        cell_address = cell.coordinate
                        formula_str = str(cell.value) if cell.value is not None else ""
                        app_logger.info(
                            f"Found formula in {sheet_name}!{cell_address}: {formula_str}",
                            context={'sheet': sheet_name, 'cell': cell_address, 'formula': formula_str}
                        )
                        extracted_formulas_raw.append({
                            "file_path": file_path,
                            "sheet_name": sheet_name,
                            "cell_address": cell_address,
                            "formula_string": formula_str
                        })
    except Exception as e:
        app_logger.error(f"Error parsing Excel file {file_path}: {e}", context={'file': file_path, 'error': str(e)})

    return extracted_formulas_raw 