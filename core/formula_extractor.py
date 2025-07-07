import os
from typing import List, Dict, Any, Optional
import re
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import keyword
import builtins
from formulas import Parser, ExcelModel
from utils.logger import AppLogger # Import the custom logger
from .excel_functions import register_custom_excel_functions # Import custom function registration

from .excel_parser import parse_excel_file, extract_named_ranges
from .csv_tsv_parser import parse_csv_tsv_file

# Instantiate the custom logger
app_logger = AppLogger(__name__)

# Register custom Excel functions at startup
register_custom_excel_functions()

class FormulaInfo:
    """Represents extracted formula information."""
    def __init__(self, file_path: str, sheet_name: str, cell_address: str, formula_string: str, 
                 variable_name: Optional[str] = None, python_formula_string: Optional[str] = None, 
                 python_code_snippet: Optional[str] = None, error: Optional[str] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.cell_address = cell_address
        self.formula_string = formula_string
        self.variable_name = variable_name
        self.python_formula_string = python_formula_string
        self.python_code_snippet = python_code_snippet
        self.error = error

    def __repr__(self):
        return f"FormulaInfo(Sheet='{self.sheet_name}', Cell='{self.cell_address}', Formula='{self.formula_string}', VariableName={self.variable_name}, PythonFormula={self.python_formula_string}, Error={self.error})"

def _sanitize_base_name(name: str) -> str:
    """
    Sanitizes a string to be a valid Python variable base name, checking for keywords/built-ins.
    """
    sanitized_name = name.lower().replace(" ", "_").replace("-", "_")
    sanitized_name = re.sub(r'[^a-z0-9_]', '', sanitized_name)
    sanitized_name = sanitized_name.strip("_")

    if not sanitized_name:
        app_logger.warning("Sanitized name is empty, using 'unnamed_field'.", context={'original_name': name})
        sanitized_name = "unnamed_field"

    if sanitized_name[0].isdigit():
        app_logger.info(f"Prepending 'col_' to numerical header: {sanitized_name}", context={'original_name': name})
        sanitized_name = "col_" + sanitized_name

    original_name = sanitized_name
    counter = 1
    while keyword.iskeyword(sanitized_name) or hasattr(builtins, sanitized_name):
        app_logger.info(f"Renaming keyword/builtin variable: {sanitized_name} to {original_name}_{counter}", context={'original_name': original_name, 'counter': counter})
        sanitized_name = f"{original_name}_{counter}"
        counter += 1
    return sanitized_name

def _generate_python_variable_name(base_name: str, existing_names: set) -> str:
    """
    Generates a unique Python-compliant variable name from a base name,
    ensuring uniqueness within the set of existing_names.
    """
    sanitized_name = base_name
    counter = 1

    while sanitized_name in existing_names:
        app_logger.info(f"Renaming duplicate variable: {sanitized_name} to {base_name}_{counter}", context={'original_name': base_name, 'counter': counter})
        sanitized_name = f"{base_name}_{counter}"
        counter += 1

    existing_names.add(sanitized_name) # Add the new unique name to the set
    return sanitized_name

def _generate_cell_address_variable_name(cell_address: str, existing_names: set) -> str:
    """
    Generates a unique Python-compliant variable name from a cell address (e.g., "A1" -> "cell_A1").
    Ensures the name is unique within the set of existing_names.
    """
    base_name = _sanitize_base_name(f"cell_{cell_address.replace('$', '')}")
    return _generate_python_variable_name(base_name, existing_names)

def extract_and_sanitize_headers(file_path: str) -> Dict[str, Dict[int, str]]:
    """
    Extracts and sanitizes headers from the first row of each sheet in an Excel file.
    Handles merged cells and returns a dictionary of sheet names to a dictionary
    mapping column index to unique, sanitized header name.
    """
    all_sanitized_headers: Dict[str, Dict[int, str]] = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True) # data_only=True to get values, not formulas

        for sheet_name in workbook.sheetnames:
            sheet: Worksheet = workbook[sheet_name]
            app_logger.info(
                f"Extracting headers from sheet: {sheet_name} in {os.path.basename(file_path)}",
                context={'sheet': sheet_name, 'file': os.path.basename(file_path)}
            )

            sanitized_headers_for_sheet: Dict[int, str] = {}
            used_names_in_sheet = set() # Track names used in this sheet for uniqueness within the sheet

            # Iterate through the first row (row_idx = 1)
            for col_idx in range(1, sheet.max_column + 1):
                cell: Cell = sheet.cell(row=1, column=col_idx) # type: ignore

                header_value = ""
                if cell.value is not None:
                    header_value = str(cell.value)
                
                # Sanitize the header value to get a base name
                base_header_name = _sanitize_base_name(header_value)

                # Generate a unique, Python-compliant name for the header within this sheet
                unique_header_name = _generate_python_variable_name(base_header_name, used_names_in_sheet)
                sanitized_headers_for_sheet[col_idx] = unique_header_name
            all_sanitized_headers[sheet_name] = sanitized_headers_for_sheet
    except Exception as e:
        app_logger.error(f"Error extracting headers from {file_path}: {e}", context={'file': file_path, 'error': str(e)})

    return all_sanitized_headers

def extract_formulas_from_file(file_path: str) -> List[FormulaInfo]:
    """
    Extracts formula information from a given file, supporting Excel, CSV, and TSV formats.
    Assigns a unique variable name to each extracted formula using fallback logic.

    Args:
        file_path (str): The path to the file.
    Returns:
        List[FormulaInfo]: A list of FormulaInfo objects containing extracted formula information.
    """
    extracted_formulas: List[FormulaInfo] = []
    filename = os.path.basename(file_path)
    file_extension = os.path.splitext(filename)[1].lower()
    
    # This set will track all unique variable names generated across the entire file processing.
    all_used_variable_names = set()

    if file_extension in ['.xls', '.xlsx']:
        app_logger.info(f"Detecting Excel file: {filename}", context={'file': filename})
        
        # First, extract and incorporate all header names into the global uniqueness set
        sanitized_headers_map = extract_and_sanitize_headers(file_path)
        for sheet_name, headers_by_col in sanitized_headers_map.items():
            for col_idx, header_name in headers_by_col.items():
                # Add to global uniqueness set just to reserve the name if it's used as a header
                _generate_python_variable_name(header_name, all_used_variable_names)

        # Extract and process named ranges
        named_ranges_dict = extract_named_ranges(file_path)
        cell_address_to_named_range_name: Dict[str, str] = {}
        for name_raw, address_in_excel_format in named_ranges_dict.items():
            sanitized_name = _sanitize_base_name(name_raw)
            unique_name = _generate_python_variable_name(sanitized_name, all_used_variable_names)
            
            address_parts = address_in_excel_format.split('!')
            sheet_name_for_lookup = None
            cell_part_for_lookup = ""

            if len(address_parts) == 2:
                # Remove quotes from sheet name if present, and convert to uppercase
                sheet_name_for_lookup = address_parts[0].upper().strip("'").strip('"')
                cell_part_for_lookup = address_parts[1].replace('$', '').upper()
            else: 
                # Assuming global or current sheet if no '!'
                cell_part_for_lookup = address_parts[0].replace('$', '').upper()

            # Only map if it's a single cell reference (not a range)
            if ':' not in cell_part_for_lookup:
                lookup_key = ""
                if sheet_name_for_lookup:
                    lookup_key = f"{sheet_name_for_lookup}!{cell_part_for_lookup}"
                else:
                    # For global named ranges, the lookup key will just be the cell address
                    lookup_key = cell_part_for_lookup
                cell_address_to_named_range_name[lookup_key] = unique_name
        
        raw_excel_formulas_data = parse_excel_file(file_path)
        for formula_data in raw_excel_formulas_data:
            formula_info = FormulaInfo(
                file_path=formula_data["file_path"],
                sheet_name=formula_data["sheet_name"],
                cell_address=formula_data["cell_address"],
                formula_string=formula_data["formula_string"]
            )

            # Determine variable name: prioritize named ranges
            current_sheet_name_upper = formula_info.sheet_name.upper()
            current_cell_address_normalized = formula_info.cell_address.replace('$', '').upper()
            
            potential_named_range_var_name = None

            # Try sheet-scoped lookup
            sheet_scoped_lookup_key = f"{current_sheet_name_upper}!{current_cell_address_normalized}"
            if sheet_scoped_lookup_key in cell_address_to_named_range_name:
                potential_named_range_var_name = cell_address_to_named_range_name[sheet_scoped_lookup_key]
            else:
                # Try global lookup (if the named range itself didn't have a sheet prefix)
                if current_cell_address_normalized in cell_address_to_named_range_name:
                    potential_named_range_var_name = cell_address_to_named_range_name[current_cell_address_normalized]

            if potential_named_range_var_name:
                formula_info.variable_name = potential_named_range_var_name
            else:
                # Try to get variable name from header if available
                # Convert cell address to column index (e.g., 'A1' -> 1, 'B5' -> 2)
                col_letter = ''.join(filter(str.isalpha, formula_info.cell_address)).upper()
                col_index = 0
                for i, char in enumerate(col_letter):
                    col_index = col_index * 26 + (ord(char) - ord('A') + 1)
                
                sheet_headers = sanitized_headers_map.get(formula_info.sheet_name, {})
                header_var_name = sheet_headers.get(col_index)

                if header_var_name:
                    formula_info.variable_name = _generate_python_variable_name(header_var_name, all_used_variable_names)
                else:
                    # Fallback to cell address variable name
                    formula_info.variable_name = _generate_cell_address_variable_name(
                        formula_info.cell_address, all_used_variable_names
                    )
            try:
                parser = Parser()
                compiled_formula = parser.ast(formula_info.formula_string)[1].compile()
                formula_info.python_formula_string = str(compiled_formula)

                # Generate a more complete Python code snippet
                code_lines = [
                    "import math",
                    "from formulas import ExcelError, Array",
                    "from datetime import datetime",
                    "# Custom Excel functions (if any) go here or are imported",
                    "# Example: from your_custom_functions_module import XLOOKUP_custom",
                    ""
                ]

                # Determine inputs for the compiled formula
                formula_inputs = []
                if hasattr(compiled_formula, 'inputs') and compiled_formula.inputs is not None:
                    try:
                        formula_inputs = [str(inp) for inp in compiled_formula.inputs]
                    except TypeError as e: # Catch if it's not directly iterable
                        app_logger.warning(
                            f"Compiled formula inputs are not directly iterable for formula: {formula_info.formula_string} - {e}",
                            context={'file': file_path, 'sheet': formula_info.sheet_name, 'cell': formula_info.cell_address}
                        )
                else:
                    app_logger.warning(
                        f"Compiled formula has no detectable inputs (or inputs is None) for formula: {formula_info.formula_string}",
                        context={'file': file_path, 'sheet': formula_info.sheet_name, 'cell': formula_info.cell_address}
                    )
                
                # Map cell references to generated Python variable names using the established hierarchy
                input_variable_names = []
                for cell_ref_raw in formula_inputs:
                    cell_ref = cell_ref_raw.replace("'", "").replace("$", "").upper() # Normalize cell reference
                    
                    # 1. Try Named Range lookup for input cell reference
                    input_var_name = None
                    input_sheet_name = None
                    input_cell_address = cell_ref

                    if '!' in cell_ref:
                        parts = cell_ref.split('!')
                        input_sheet_name = parts[0]
                        input_cell_address = parts[1]
                        # Try sheet-scoped named range
                        lookup_key = f"{input_sheet_name}!{input_cell_address}"
                        if lookup_key in cell_address_to_named_range_name:
                            input_var_name = cell_address_to_named_range_name[lookup_key]
                    
                    if input_var_name is None:
                        # Try global named range lookup
                        if input_cell_address in cell_address_to_named_range_name:
                            input_var_name = cell_address_to_named_range_name[input_cell_address]
                    
                    if input_var_name is None:
                        # 2. Try Header lookup for input cell reference
                        if input_sheet_name is None: # Assume current sheet if no sheet specified in input cell_ref
                            input_sheet_name = formula_info.sheet_name.upper()

                        # Convert cell address to column index for header lookup
                        input_col_letter = ''.join(filter(str.isalpha, input_cell_address)).upper()
                        input_col_index = 0
                        for i, char in enumerate(input_col_letter):
                            input_col_index = input_col_index * 26 + (ord(char) - ord('A') + 1)
                        
                        sheet_headers = sanitized_headers_map.get(input_sheet_name, {})
                        header_for_input = sheet_headers.get(input_col_index)

                        if header_for_input:
                            input_var_name = _generate_python_variable_name(header_for_input, all_used_variable_names)
                    
                    if input_var_name is None:
                        # 3. Fallback to Cell Address variable name
                        input_var_name = _generate_cell_address_variable_name(cell_ref_raw, all_used_variable_names)
                    
                    input_variable_names.append(input_var_name)
                
                # Remove duplicates while preserving order
                unique_input_variable_names = []
                seen = set()
                for name in input_variable_names:
                    if name not in seen:
                        unique_input_variable_names.append(name)
                        seen.add(name)

                # Generate code lines for inputs
                for var_name in unique_input_variable_names:
                    # This is a placeholder; actual values will come from parsed data
                    code_lines.append(f"{var_name} = None # Placeholder for input value from spreadsheet")
                
                # Add the formula calculation line
                if formula_info.variable_name:
                    code_lines.append(f"{formula_info.variable_name} = {formula_info.python_formula_string}")
                else:
                    # Fallback if for some reason main variable name was not set (should not happen)
                    code_lines.append(f"cell_{formula_info.cell_address.replace('$', '')} = {formula_info.python_formula_string}")
                
                formula_info.python_code_snippet = "\n".join(code_lines)
                
                extracted_formulas.append(formula_info)

            except Exception as e:
                app_logger.error(
                    f"Error processing formula '{formula_info.formula_string}' in {formula_info.sheet_name}!{formula_info.cell_address}: {e}",
                    context={'file': file_path, 'sheet': formula_info.sheet_name, 'cell': formula_info.cell_address, 'formula': formula_info.formula_string, 'error': str(e)}
                )
                formula_info.error = str(e)
                extracted_formulas.append(formula_info)

    elif file_extension in ['.csv', '.tsv']:
        app_logger.info(f"Detecting CSV/TSV file: {filename}", context={'file': filename})
        csv_tsv_data = parse_csv_tsv_file(file_path)
        # For CSV/TSV, formulas are not natively supported, but we might look for '=' prefix
        # This part needs further refinement based on how formulas are expected in CSV/TSV
        app_logger.warning(f"CSV/TSV formula extraction is not fully implemented for dynamic variable naming based on headers or named ranges. Using cell-based naming.", context={'file': file_path})

        for row_idx, row_data in enumerate(csv_tsv_data):
            for col_idx, cell_value in enumerate(row_data):
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_address = f'{openpyxl.utils.get_column_letter(col_idx + 1)}{row_idx + 1}'
                    formula_info = FormulaInfo(
                        file_path=file_path,
                        sheet_name="Sheet1", # CSV/TSV treated as single sheet
                        cell_address=cell_address,
                        formula_string=cell_value
                    )
                    formula_info.variable_name = _generate_cell_address_variable_name(
                        formula_info.cell_address, all_used_variable_names
                    )
                    try:
                        parser = Parser()
                        compiled_formula = parser.ast(formula_info.formula_string.lstrip('='))[1].compile()
                        formula_info.python_formula_string = str(compiled_formula)

                        code_lines = [
                            "import math",
                            "from formulas import ExcelError, Array",
                            "from datetime import datetime",
                            ""
                        ]

                        formula_inputs = []
                        if hasattr(compiled_formula, 'inputs') and compiled_formula.inputs is not None:
                            try:
                                formula_inputs = [str(inp) for inp in compiled_formula.inputs]
                            except TypeError as e:
                                app_logger.warning(
                                    f"Compiled formula inputs are not directly iterable for CSV/TSV formula: {formula_info.formula_string} - {e}",
                                    context={'file': file_path, 'sheet': formula_info.sheet_name, 'cell': formula_info.cell_address}
                                )
                        
                        input_variable_names = []
                        for cell_ref in formula_inputs:
                            # For CSV/TSV, we primarily rely on cell address naming for inputs
                            input_var_name = _generate_cell_address_variable_name(cell_ref.replace("'", "").replace("!", "_"), all_used_variable_names)
                            input_variable_names.append(input_var_name)
                        
                        unique_input_variable_names = []
                        seen = set()
                        for name in input_variable_names:
                            if name not in seen:
                                unique_input_variable_names.append(name)
                                seen.add(name)

                        for var_name in unique_input_variable_names:
                            code_lines.append(f"{var_name} = None # Placeholder for input value from CSV/TSV")
                        
                        if formula_info.variable_name:
                            code_lines.append(f"{formula_info.variable_name} = {formula_info.python_formula_string}")
                        else:
                            code_lines.append(f"cell_{formula_info.cell_address.replace('$', '')} = {formula_info.python_formula_string}")
                        
                        formula_info.python_code_snippet = "\n".join(code_lines)
                        extracted_formulas.append(formula_info)
                    except Exception as e:
                        app_logger.error(
                            f"Error processing CSV/TSV formula '{formula_info.formula_string}' in {formula_info.sheet_name}!{formula_info.cell_address}: {e}",
                            context={'file': file_path, 'sheet': formula_info.sheet_name, 'cell': formula_info.cell_address, 'formula': formula_info.formula_string, 'error': str(e)}
                        )
                        formula_info.error = str(e)
                        extracted_formulas.append(formula_info)
    else:
        app_logger.error(f"Unsupported file type: {file_extension}", context={'file': file_path, 'extension': file_extension})
        # Optionally, you could append a FormulaInfo object with an error for the file itself

    return extracted_formulas

def extract_formula_data(file_path: str, file_type: str) -> List[Dict[str, str]]:
    """
    Extracts raw formula data from a given file based on its type.

    Args:
        file_path (str): The path to the input file.
        file_type (str): The type of the file ('xlsx', 'csv', or 'tsv').

    Returns:
        List[Dict[str, str]]: A list of dictionaries, each containing raw formula information.
    """
    extracted_data = []
    try:
        if file_type == 'xlsx':
            extracted_data = parse_excel_file(file_path)
        elif file_type == 'csv' or file_type == 'tsv':
            extracted_data = parse_csv_tsv_file(file_path)
        else:
            app_logger.error(f"Unsupported file type: {file_type}")
    except Exception as e:
        app_logger.error(f"Error extracting formula data from {file_path}: {e}")
    return extracted_data 